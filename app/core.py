from openpyxl import load_workbook
from datetime import date
from sys import argv
from random import choice

from app.conf import *
from app.database import Storage


class ExcelParser:
    """
    Класс отвечает за обработку Excel-файла.
    """
    def __init__(self):
        self.data = []

    def parse_excel(self):
        """
        Метод получает из параметров командной строки имя файла.
        Заполняет список data словарями, содержащими
        данные отдельных строк файла.
        """

        # Очищаем список от предыдущих данных, если бы они были
        self.data.clear()

        # Открываем файл, иначе выходим с ошибкой
        try:
            file_name = argv[1]
            workbook = load_workbook(file_name)
        except FileNotFoundError:
            print("Файл не найден.")
            exit(1)
        except Exception:
            print("Произошла ошибка")
            exit(1)

        # Выбираем текущий лист документа и пропускаем шапку
        worksheet = workbook.active
        worksheet = worksheet[TABLE_HEADER : worksheet.max_row]

        # Парсим ячейки в соответствии с настройками из conf.py
        for row in worksheet:
            item = {
                "company": row[COMPANY].value,
                "fact_qliq_1": row[FACT_QLIQ_1].value,
                "fact_qliq_2": row[FACT_QLIQ_2].value,
                "fact_qoil_1": row[FACT_QOIL_1].value,
                "fact_qoil_2": row[FACT_QOIL_2].value,
                "forecast_qliq_1": row[FORECAST_QLIQ_1].value,
                "forecast_qliq_2": row[FORECAST_QLIQ_2].value,
                "forecast_qoil_1": row[FORECAST_QOIL_1].value,
                "forecast_qoil_2": row[FORECAST_QOIL_2].value,
                "date": date(2023, 1, choice(DAYS)),
            }
            self.data.append(item)


class Parser(ExcelParser):
    """
    Класс отвечает за сохранение данных в бд и вычисление сумм
    """
    def __init__(self):
        super().__init__()
        self.db = Storage()

    def save_to_db(self):
        self.db.save_data(self.data)
        self.db.calculate_sum()


if __name__ == "__main__":
    file = Parser()
    file.parse_excel()
    file.save_to_db()
