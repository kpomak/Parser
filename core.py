from openpyxl import load_workbook
from datetime import date
from sys import argv

from conf import *


class ExcelParser:
    def parse_excel(self):
        try:
            file_name = argv[1]
            workbook = load_workbook(file_name)
        except FileNotFoundError:
            print('Файл не найден.')
            exit(1)
        except Exception:
            print('Произошла ошибка')
            exit(1)

        worksheet = workbook.active
        worksheet = worksheet[TABLE_HEADER:worksheet.max_row]

        data = []
        day_of_month = (date(2023, 1, day) for day in range(1, 32))
        for row in worksheet:
            item = {
                'company': row[COMPANY].value,
                'fact_qliq_1': row[FACT_QLIQ_1].value,
                'fact_qliq_2': row[FACT_QLIQ_2].value,
                'fact_qoil_1': row[FACT_QOIL_1].value,
                'fact_qoil_2': row[FACT_QOIL_2].value,
                'forecast_qliq_1': row[FORECAST_QLIQ_1].value,
                'forecast_qliq_2': row[FORECAST_QLIQ_2].value,
                'forecast_qoil_1': row[FORECAST_QOIL_1].value,
                'forecast_qoil_2': row[FORECAST_QOIL_2].value,
                'date': next(day_of_month),
            }
            data.append(item)


class Parser(ExcelParser):
    pass


if __name__ == "__main__":
    file = ExcelParser()
    file.parse_excel()
